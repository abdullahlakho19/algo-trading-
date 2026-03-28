"""
reporting/excel_exporter.py
─────────────────────────────────────────────────────────────────────────────
Excel Reporting Engine.
Exports complete trade history, performance metrics, and model health
to a professionally formatted Excel workbook.
─────────────────────────────────────────────────────────────────────────────
"""

import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime
from pathlib import Path
from core.logger import get_logger
from config import config

log = get_logger(__name__)

# ── Colors ────────────────────────────────────────────────────────────────────
C_DARK   = "0A0E1A"
C_GOLD   = "C9A84C"
C_CYAN   = "00D4FF"
C_GREEN  = "00C853"
C_RED    = "FF5252"
C_GRAY   = "D0D6E0"
C_WHITE  = "FFFFFF"
C_NAVY   = "0D1B2A"
C_MID    = "1E2A3A"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_WHITE, size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def _border() -> Border:
    side = Side(style="thin", color="2A3A4A")
    return Border(left=side, right=side, top=side, bottom=side)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


class ExcelExporter:
    """
    Exports trading performance to a multi-sheet Excel workbook.

    Sheets:
        1. Summary Dashboard
        2. Trade History
        3. Performance Metrics
        4. Daily P&L
        5. Open Positions
    """

    def __init__(self):
        self.output_dir = config.paths.reports
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ── Main Export ───────────────────────────────────────────────────────────
    def export(
        self,
        closed_trades:    list,
        open_positions:   list = None,
        portfolio_value:  float = 100_000.0,
        initial_capital:  float = 100_000.0,
        filename:         str = None,
    ) -> Path:
        """
        Generate full Excel report.

        Args:
            closed_trades:   List of ClosedTrade objects (from paper_simulator)
            open_positions:  List of open SimulatedPosition objects
            portfolio_value: Current portfolio equity
            initial_capital: Starting capital
            filename:        Optional output filename

        Returns:
            Path to generated Excel file
        """
        if not filename:
            ts       = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"trading_report_{ts}.xlsx"

        output_path = self.output_dir / filename
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # Remove default sheet

        # Build all sheets
        self._build_summary(wb, closed_trades, portfolio_value, initial_capital)
        self._build_trade_history(wb, closed_trades)
        self._build_performance(wb, closed_trades)
        self._build_daily_pnl(wb, closed_trades)
        if open_positions:
            self._build_open_positions(wb, open_positions)

        wb.save(output_path)
        log.info(f"Excel report exported: {output_path}")
        return output_path

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    def _build_summary(
        self, wb, trades: list, equity: float, initial: float
    ) -> None:
        ws = wb.create_sheet("📊 Summary")
        ws.sheet_view.showGridLines = False

        # Header
        ws.merge_cells("A1:H1")
        ws["A1"] = "INSTITUTIONAL TRADING AGENT — PERFORMANCE SUMMARY"
        ws["A1"].font      = _font(bold=True, size=14, color=C_GOLD)
        ws["A1"].fill      = _fill(C_DARK)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws["A2"].font      = _font(size=9, color=C_GRAY, italic=True)
        ws["A2"].fill      = _fill(C_NAVY)
        ws["A2"].alignment = _center()

        # Metrics
        if trades:
            wins   = [t for t in trades if t.outcome == "win"]
            losses = [t for t in trades if t.outcome == "loss"]
            win_rate = len(wins) / len(trades)
            total_pnl = sum(t.pnl for t in trades)
            total_pnl_pct = (equity - initial) / initial

            gross_profit = sum(t.pnl for t in wins) if wins else 0
            gross_loss   = abs(sum(t.pnl for t in losses)) if losses else 1
            profit_factor = gross_profit / gross_loss

            pnl_series = pd.Series([t.pnl for t in trades])
            sharpe = (pnl_series.mean() / pnl_series.std() * (252 ** 0.5)
                      ) if len(pnl_series) > 1 else 0

        else:
            win_rate = total_pnl = total_pnl_pct = profit_factor = sharpe = 0
            wins = losses = []

        metrics = [
            ("", "", "", "", "", "", "", ""),
            ("METRIC", "VALUE", "TARGET", "STATUS", "", "METRIC", "VALUE", "STATUS"),
            ("Total Trades",     len(trades),              "—",    "—",          "", "Portfolio Value",   f"${equity:,.2f}",          "—"),
            ("Win Rate",         f"{win_rate:.1%}",        "75%+", "✅" if win_rate >= 0.75 else "⚠️", "", "Total P&L",          f"${total_pnl:,.2f}",       "✅" if total_pnl > 0 else "❌"),
            ("Profit Factor",    f"{profit_factor:.2f}",   "2.0+", "✅" if profit_factor >= 2 else "⚠️", "", "P&L %",             f"{total_pnl_pct:.2%}",    "✅" if total_pnl_pct > 0 else "❌"),
            ("Sharpe Ratio",     f"{sharpe:.2f}",          "2.0+", "✅" if sharpe >= 2 else "⚠️",       "", "Initial Capital",    f"${initial:,.2f}",         "—"),
            ("Total Wins",       len(wins),                "—",    "—",          "", "Mode",              "PAPER TRADING",            "✅"),
            ("Total Losses",     len(losses),              "—",    "—",          "", "Last Update",        datetime.utcnow().strftime("%H:%M UTC"), "—"),
        ]

        row_start = 4
        for r_idx, row in enumerate(metrics, start=row_start):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == row_start + 1:   # Header row
                    cell.fill = _fill(C_GOLD)
                    cell.font = _font(bold=True, color=C_DARK, size=9)
                elif c_idx in (1, 6):         # Label columns
                    cell.fill = _fill(C_MID)
                    cell.font = _font(bold=True, color=C_GOLD, size=9)
                elif c_idx in (4, 8):         # Status columns
                    cell.fill = _fill(C_NAVY)
                    cell.font = _font(size=11)
                else:
                    cell.fill = _fill(C_NAVY)
                    cell.font = _font(size=9, color=C_WHITE)
                cell.alignment = _center()
                cell.border    = _border()

        # Column widths
        widths = [18, 14, 10, 8, 3, 20, 16, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Trade History ────────────────────────────────────────────────
    def _build_trade_history(self, wb, trades: list) -> None:
        ws = wb.create_sheet("📋 Trade History")
        ws.sheet_view.showGridLines = False

        headers = [
            "Trade ID", "Symbol", "Market", "Direction", "Qty",
            "Entry Price", "Exit Price", "Stop Loss", "Take Profit",
            "P&L ($)", "P&L (%)", "Outcome", "R/R",
            "Duration (min)", "Exit Reason", "Opened At", "Closed At"
        ]

        # Header row
        ws.row_dimensions[1].height = 30
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill      = _fill(C_GOLD)
            cell.font      = _font(bold=True, color=C_DARK, size=9)
            cell.alignment = _center()
            cell.border    = _border()

        # Data rows
        for r, trade in enumerate(trades, start=2):
            rr = abs(trade.take_profit - trade.entry_price) / abs(trade.entry_price - trade.stop_loss) if trade.stop_loss != trade.entry_price else 0
            row_data = [
                trade.trade_id,
                trade.symbol,
                trade.market.upper(),
                trade.direction.upper(),
                round(trade.qty, 4),
                round(trade.entry_price, 6),
                round(trade.exit_price, 6),
                round(trade.stop_loss, 6),
                round(trade.take_profit, 6),
                round(trade.pnl, 2),
                round(trade.pnl_pct * 100, 2),
                trade.outcome.upper(),
                round(rr, 2),
                round(trade.duration_min, 1),
                trade.exit_reason.replace("_", " ").upper(),
                trade.opened_at.strftime("%Y-%m-%d %H:%M"),
                trade.closed_at.strftime("%Y-%m-%d %H:%M"),
            ]

            bg = C_MID if r % 2 == 0 else C_NAVY

            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill      = _fill(bg)
                cell.alignment = _center()
                cell.border    = _border()
                cell.font      = _font(size=8.5)

                # Colour P&L column
                if c == 10:
                    cell.font = _font(
                        bold=True, size=8.5,
                        color=C_GREEN if (val or 0) > 0 else C_RED
                    )
                if c == 12:
                    cell.font = _font(
                        bold=True, size=8.5,
                        color=C_GREEN if val == "WIN" else C_RED
                    )

        # Column widths
        col_widths = [10, 8, 8, 9, 8, 12, 12, 12, 12, 10, 8, 9, 6, 13, 12, 16, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Performance Metrics ──────────────────────────────────────────
    def _build_performance(self, wb, trades: list) -> None:
        ws = wb.create_sheet("📈 Performance")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:D1")
        ws["A1"] = "PERFORMANCE METRICS"
        ws["A1"].font = _font(bold=True, size=12, color=C_GOLD)
        ws["A1"].fill = _fill(C_DARK)
        ws["A1"].alignment = _center()

        if not trades:
            ws["A2"] = "No closed trades yet."
            return

        wins   = [t for t in trades if t.outcome == "win"]
        losses = [t for t in trades if t.outcome == "loss"]
        pnl_series = pd.Series([t.pnl for t in trades])

        metrics = [
            ("METRIC", "VALUE", "BENCHMARK", "PASS/FAIL"),
            ("Total Trades",          len(trades),                          "—",        "—"),
            ("Win Rate",              f"{len(wins)/len(trades):.2%}",       "≥ 75%",    "✅" if len(wins)/len(trades) >= 0.75 else "❌"),
            ("Loss Rate",             f"{len(losses)/len(trades):.2%}",     "≤ 25%",    "✅" if len(losses)/len(trades) <= 0.25 else "❌"),
            ("Total P&L",             f"${sum(t.pnl for t in trades):,.2f}", "Positive", "✅" if sum(t.pnl for t in trades) > 0 else "❌"),
            ("Gross Profit",          f"${sum(t.pnl for t in wins):,.2f}",  "—",        "—"),
            ("Gross Loss",            f"${abs(sum(t.pnl for t in losses)):,.2f}", "—",  "—"),
            ("Avg Win",               f"${pnl_series[pnl_series > 0].mean():.2f}" if any(t.pnl > 0 for t in trades) else "$0", "—", "—"),
            ("Avg Loss",              f"${abs(pnl_series[pnl_series < 0].mean()):.2f}" if any(t.pnl < 0 for t in trades) else "$0", "—", "—"),
            ("Profit Factor",         f"{abs(sum(t.pnl for t in wins))/(abs(sum(t.pnl for t in losses))+0.01):.2f}", "≥ 2.0", "✅" if abs(sum(t.pnl for t in wins))/(abs(sum(t.pnl for t in losses))+0.01) >= 2 else "❌"),
            ("Sharpe Ratio",          f"{pnl_series.mean()/pnl_series.std()*(252**0.5):.2f}" if len(pnl_series) > 1 else "—", "≥ 2.0", "—"),
            ("Max Single Win",        f"${max((t.pnl for t in wins), default=0):,.2f}",  "—", "—"),
            ("Max Single Loss",       f"${min((t.pnl for t in losses), default=0):,.2f}", "—", "—"),
            ("Avg Duration (min)",    f"{sum(t.duration_min for t in trades)/len(trades):.1f}", "—", "—"),
        ]

        for r, row in enumerate(metrics, start=2):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                if r == 2:
                    cell.fill = _fill(C_GOLD)
                    cell.font = _font(bold=True, color=C_DARK)
                elif c == 1:
                    cell.fill = _fill(C_MID)
                    cell.font = _font(bold=True, color=C_GOLD, size=9)
                else:
                    cell.fill = _fill(C_NAVY if r % 2 == 0 else C_MID)
                    cell.font = _font(size=9)
                cell.alignment = _center()
                cell.border    = _border()

        for i, w in enumerate([22, 16, 12, 10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: Daily P&L ────────────────────────────────────────────────────
    def _build_daily_pnl(self, wb, trades: list) -> None:
        ws = wb.create_sheet("📅 Daily P&L")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:E1")
        ws["A1"] = "DAILY P&L BREAKDOWN"
        ws["A1"].font = _font(bold=True, size=12, color=C_GOLD)
        ws["A1"].fill = _fill(C_DARK)
        ws["A1"].alignment = _center()

        if not trades:
            return

        df = pd.DataFrame([{
            "date": t.closed_at.date(),
            "pnl":  t.pnl,
            "outcome": t.outcome,
        } for t in trades])

        daily = df.groupby("date").agg(
            total_pnl=("pnl", "sum"),
            trades=("pnl", "count"),
            wins=("outcome", lambda x: (x == "win").sum()),
        ).reset_index()
        daily["cumulative_pnl"] = daily["total_pnl"].cumsum()

        headers = ["Date", "Daily P&L", "Trades", "Wins", "Cumulative P&L"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.fill      = _fill(C_GOLD)
            cell.font      = _font(bold=True, color=C_DARK, size=9)
            cell.alignment = _center()
            cell.border    = _border()

        for r, row in daily.iterrows():
            bg = C_MID if r % 2 == 0 else C_NAVY
            data = [
                str(row["date"]),
                round(row["total_pnl"], 2),
                int(row["trades"]),
                int(row["wins"]),
                round(row["cumulative_pnl"], 2),
            ]
            for c, val in enumerate(data, 1):
                cell = ws.cell(row=r + 3, column=c, value=val)
                cell.fill      = _fill(bg)
                cell.font      = _font(size=9, color=C_GREEN if isinstance(val, float) and val > 0 else C_RED if isinstance(val, float) and val < 0 else C_WHITE)
                cell.alignment = _center()
                cell.border    = _border()

        for i, w in enumerate([12, 12, 8, 8, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 5: Open Positions ───────────────────────────────────────────────
    def _build_open_positions(self, wb, positions: list) -> None:
        ws = wb.create_sheet("🔓 Open Positions")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:G1")
        ws["A1"] = "CURRENT OPEN POSITIONS"
        ws["A1"].font = _font(bold=True, size=12, color=C_GOLD)
        ws["A1"].fill = _fill(C_DARK)
        ws["A1"].alignment = _center()

        headers = ["Symbol", "Direction", "Qty", "Entry Price", "Current Price", "Stop Loss", "Unrealised P&L"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.fill      = _fill(C_GOLD)
            cell.font      = _font(bold=True, color=C_DARK, size=9)
            cell.alignment = _center()
            cell.border    = _border()

        for r, pos in enumerate(positions, start=3):
            data = [
                pos.symbol, pos.direction.upper(),
                round(pos.qty, 4), round(pos.entry_price, 6),
                round(pos.current_price, 6), round(pos.stop_loss, 6),
                round(pos.unrealised_pnl, 2),
            ]
            for c, val in enumerate(data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill      = _fill(C_MID if r % 2 == 0 else C_NAVY)
                cell.font      = _font(size=9)
                cell.alignment = _center()
                cell.border    = _border()

        for i, w in enumerate([10, 10, 8, 14, 14, 12, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w


# Singleton
excel_exporter = ExcelExporter()